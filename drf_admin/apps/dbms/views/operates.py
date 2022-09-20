from django.views.decorators.csrf import csrf_exempt
from rest_framework.decorators import api_view
from dbms.engines.operateLogs import operateLogs
from dbms.models import OperateLogs, DBServerConfig, Audits, CheckContent
from rest_framework.response import Response
from django_filters.rest_framework.filterset import FilterSet
from django_filters import filters
from rest_framework.views import APIView
from rest_framework.generics import ListCreateAPIView, RetrieveAPIView
from dbms.serializers.operates import OperateLogsSerializer
from django_filters.rest_framework import DjangoFilterBackend
from dbms.engines.goInception import GoInceptionEngine
from rest_framework.filters import SearchFilter
from drf_yasg.utils import swagger_auto_schema
from drf_yasg import openapi
from dbms.serializers.audits import AuditsSerializer
from drf_admin.utils.views import AdminViewSet
from rest_framework import status
from celery_tasks.dingding.tasks import send_dingding_msg
from django.views.decorators.http import require_POST
from django.http import HttpResponse
from openpyxl import Workbook
from dbms.engines.mysql import MysqlEngine
import os, json, re
import drf_admin.common.vars as var


access_token = var.sql_token


def check(data, conn):
    """
    检查提交sql；
    """
    db = data["db"]
    # db_name是list，检查时，检查一个库即可
    db_name = data["execute_db_name"][0]
    sql = data["operate_sql"]
    # 调用goInception中检查方法
    result = GoInceptionEngine.check(db, db_name, sql, conn)
    result["dbName"] = db_name
    if "error" in result:
        return result
    # 将检查结果放入数据库
    if result["errorCount"] > 0:
        CheckContent.objects.create(sql_content=sql, status=0)
        # 存在异常sql时，存入操作日志
        operate_log = {}
        operate_log["db_env"] = data["environment"]
        operate_log["db_name"] = db_name
        operate_log["performer"] = data["performer"]
        operate_log["sql"] = sql
        operate_log["status"] = 0
        operate_log["sprint"] = data["sprint"]
        # 存在error代表check出现异常
        if "error" in result:
            operate_log["error_message"] = result["error"]
            return Response(status=status.HTTP_400_BAD_REQUEST, data={"error": result["error"]})
        # result["errorCount"] > 0代表存在异常sql
        if result["errorCount"] > 0:
            data["errors"] = result
            operate_log["error_message"] = "sql检查不通过！"
            operateLogs(operate_log)
    else:
        CheckContent.objects.create(sql_content=sql, status=1)
    return result


class DatabasesView(APIView):
    """
    get:
    数据库操作--获取全部数据库

    获取信息详情, status: 201(成功), return: 数据库表
    post:
    数据库操作--执行sql

    获取信息详情, status: 201(成功), return: 执行结果
    """
    def get(self, request, pk):
        """获取所有数据库"""
        obj = MysqlEngine(pk)
        # 获取某个环境的数据
        all_db_list = obj.get_all_db()
        if isinstance(all_db_list, dict):
            return Response(status=status.HTTP_400_BAD_REQUEST, data={"error": all_db_list["error"]})
        if request.query_params:
            # tenant用来过滤租户库
            if "tenant" not in request.query_params:
                return Response(status=status.HTTP_400_BAD_REQUEST, data={"error": "请求参数为tenant"})
            for all_db_i in all_db_list[:]:
                if not all_db_i["Database"].startswith("guozhi_tenant_") or all_db_i["Database"] == "guozhi_tenant_1":
                    all_db_list.remove(all_db_i)
        return Response(all_db_list)

    def post(self, request):
        """执行sql，并记录到日志"""
        if ["db", "execute_db_name", "operate_sql"].sort() != list(request.data.keys()).sort():
            return Response(status=status.HTTP_400_BAD_REQUEST, data={"error": "请求参数缺失！"})
        if not request.data["db"] or not request.data["execute_db_name"] or not request.data["operate_sql"]:
            return Response(status=status.HTTP_400_BAD_REQUEST, data={"error": "参数不能为空！"})
        db_names = request.data["execute_db_name"]
        sql = request.data["operate_sql"]
        request.data["sprint"] = request.data["sprint"] if "sprint" in request.data else None
        db = request.data["db"]
        obj = MysqlEngine(db)
        # 获取对应数据库连接信息
        base_data = obj.basic_info()
        if "error" in base_data:
            return Response(status=status.HTTP_400_BAD_REQUEST, data={"error": base_data["error"]})
        request.data["performer"] = request.user.get_username()
        request.data["environment"] = base_data["environment"]
        # 存储基本信息，用来存操作日志
        operate_log = {}
        # 存放返回数据
        data = {"success": [], "errors": None}
        # 建立连接
        conn = GoInceptionEngine().get_connection()
        if isinstance(conn, dict):
            return Response(status=status.HTTP_400_BAD_REQUEST, data={"error": "建立连接失败，失败原因：%s" % conn["error"]})
        # 检查sql是否验证通过
        result = check(request.data, conn)
        if result["errorCount"] > 0:
            # 发送钉钉消息
            send_dingding_msg.delay(access_token, "link", {
                "messageUrl": f"http://{request.META['HTTP_HOST']}/api/dbms/audits/?page=1&size=10&status=0",
                "picUrl": "https://img0.baidu.com/it/u=3821549314,1624213915&fm=253&fmt=auto&app=138&f=JPEG?w=773&h=500",
                "title": "SQL检查",
                "text": "sql检查失败！"
            })
            data["errors"] = result
            return Response(status=status.HTTP_400_BAD_REQUEST, data={"error": "sql检查不通过！", "errorData": data})

        # 循环调用goInception中执行方法，操作多个数据库
        i = 0
        for db_name in db_names:
            operate_log.pop("status", None)
            result = GoInceptionEngine.execute(db, db_name, sql, conn)
            operate_log["db_env"] = base_data["environment"]
            operate_log["db_name"] = db_name
            operate_log["performer"] = request.user.get_username()
            operate_log["sql"] = sql
            operate_log["sprint"] = request.data["sprint"]
            if "error" in result:
                # 发送钉钉消息
                send_dingding_msg.delay(access_token, "link", {
                    "messageUrl": f"http://{request.META['HTTP_HOST']}/api/dbms/audits/?page=1&size=10&status=0",
                    "picUrl": "https://img0.baidu.com/it/u=3821549314,1624213915&fm=253&fmt=auto&app=138&f=JPEG?w=773&h=500",
                    "title": "SQL执行",
                    "text": f"sql执行失败：成功：{i}个；失败数据库：{db_name}！"
                })
                return Response(status=status.HTTP_400_BAD_REQUEST, data={"error": result["error"]})
            result["dbName"] = db_name
            # 将执行结果放入数据库
            if result["errorCount"] > 0:
                data["errors"] = result
                operate_log["status"] = 0
                for row in result["rows"]:
                    # goInception中0代表成功，1代表warn，2代表error
                    if row["error_level"] == 2:
                        operate_log["error_message"] = row["error_message"]
                        break
                # 存入日志
                operateLogs(operate_log)
                # 发送钉钉消息
                send_dingding_msg.delay(access_token, "link", {
                    "messageUrl": f"http://{request.META['HTTP_HOST']}/api/dbms/audits/?page=1&size=10&status=0",
                    "picUrl": "https://img0.baidu.com/it/u=3821549314,1624213915&fm=253&fmt=auto&app=138&f=JPEG?w=773&h=500",
                    "title": "SQL执行",
                    "text": f"sql执行失败：成功：{i}个；失败数据库：{db_name}！"
                })
                return Response(status=status.HTTP_400_BAD_REQUEST, data={"error": "sql执行不通过！", "errorData": data})
            data["success"].append(result)
            i += 1
            # 存入操作日志
            operateLogs(operate_log)
        conn.close()
        send_dingding_msg.delay(access_token, "text", "全部SQL执行成功！")
        return Response(data=data)


class MyFilterSet(FilterSet):
    # 自定义过滤字段
    operational_data = filters.CharFilter(field_name="operate_sql", lookup_expr="icontains")
    status = filters.CharFilter(field_name="status", lookup_expr="icontains")

    class Meta:
        model = OperateLogs
        fields = ["operate_sql", "status"]


class OperateLogsPkView(RetrieveAPIView):
    """
    get:
    数据库执行日志--详情信息

    获取信息详情, status: 201(成功), return: 日志详情
    """
    # 获取某个执行日志详情
    queryset = OperateLogs.objects.order_by("-create_time")
    serializer_class = OperateLogsSerializer


class OperateLogsView(ListCreateAPIView):
    """
    get:
    数据库执行日志--列表

    日志列表, status: 201(成功), return: 列表
    post:
    数据库执行日志--创建

    创建日志, status: 201(成功), return: 日志信息
    """
    # 创建和获取执行日志
    queryset = OperateLogs.objects.order_by("-create_time")
    serializer_class = OperateLogsSerializer
    # 自定义过滤字段
    filter_backends = (DjangoFilterBackend, SearchFilter)
    filter_fields = ['status', "env"]
    search_fields = ("operate_sql", "db_name", "sprint")

    def create(self, data, **kwargs):
        serializer = OperateLogsSerializer(data=data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        return Response(serializer.data)


class AuditsViewSet(AdminViewSet):
    """
    list:
    审核--列表

    审核列表, status: 201(成功), return: 列表
    create:
    审核--创建

    创建审核, status: 201(成功), return: 添加信息
    update:
    审核--更新，支持局部

    进行审核操作, status: 201(成功), return: 更新后信息
    retrieve:
    审核--详情信息

    审核信息, status: 201(成功), return: 审核数据信息
    multiple_update:
    审核--批量更新，支持局部（目前已废弃）

    审核信息, status: 201(成功), return: 更新后数据信息
    multiple_delete:
    审核--批量删除

    审核信息, status: 201(成功), return: null
    destroy:
    审核--删除

    审核信息, status: 201(成功), return: null
    """
    queryset = Audits.objects.order_by("-update_time")
    serializer_class = AuditsSerializer
    # 自定义过滤字段
    filter_backends = (DjangoFilterBackend, SearchFilter)
    filter_fields = ["status"]
    search_fields = ("sprint",)

    def create(self, request, *args, **kwargs):
        user = request.user.get_username()
        request.data["user"] = user
        request.data["performer"] = user
        db = request.data["db"]
        obj = MysqlEngine(db)
        base_data = obj.basic_info()
        request.data["environment"] = base_data["environment"]
        # 获取对应数据库连接信息
        conn = GoInceptionEngine().get_connection()
        if isinstance(conn, dict):
            return Response(status=status.HTTP_400_BAD_REQUEST, data={"error": "建立连接失败，失败原因：%s" % conn["error"]})
        # 检查sql是否验证通过
        result = check(request.data, conn)
        conn.close()
        # 代表执行异常
        if "error" in result:
            return Response(status=status.HTTP_400_BAD_REQUEST, data={"error": result["error"]})
        # result["errorCount"] > 0代表存在检查异常语句
        if result["errorCount"] > 0:
            return Response(status=status.HTTP_400_BAD_REQUEST, data={"error": "sql检查不通过！", "errorData": result})
        request.data["execute_db_name"] = json.dumps(request.data["execute_db_name"])
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        send_dingding_msg.delay(access_token, "link", {
            "messageUrl": f"http://{request.META['HTTP_HOST']}/api/dbms/audits/?page=1&size=10&status=0",
            "picUrl": "https://img0.baidu.com/it/u=3821549314,1624213915&fm=253&fmt=auto&app=138&f=JPEG?w=773&h=500",
            "title": "SQL审核",
            "text": "这是上线需要审核的sql"
        })
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    # def multiple_update(self, request, *args, **kwargs):
    #     auditor = request.user.get_username()
    #     request.data["auditor"] = auditor
    #     return super().multiple_update(request, *args, **kwargs)

    def update(self, request, pk=None, *args, **kwargs):
        auditor = request.user.get_username()
        request.data["auditor"] = auditor
        kwargs['partial'] = True
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        # 根据审核状态发送钉钉消息
        if request.data["status"] == "1":
            send_dingding_msg.delay(access_token, "text", "sql已审核通过，请观察操作日志，是否执行成功！")
        elif request.data["status"] == "2":
            send_dingding_msg.delay(access_token, "text", f"sql被驳回，驳回原因：{request.data['reason']}！")
        return Response(serializer.data)


re_params = openapi.Parameter('file_url', openapi.IN_QUERY, description="请求参数：导出路径，非必填，默认桌面", type=openapi.TYPE_STRING)
# 导出查询数据
@swagger_auto_schema(
    method='post',
    operation_summary='导出查询数据',
    manual_parameters=[re_params],
    request_body=openapi.Schema(
        type=openapi.TYPE_OBJECT,
        properties={
            "data": openapi.Schema(title="此请求体用查询的响应data，格式为{'data': data}", type=openapi.TYPE_STRING),
        }
    ),
    responses={200: "ok"}
)
@require_POST
@csrf_exempt
@api_view(["POST"])
def export_excel(request):
    """
    post:
    导出查询数据；

    请求params：?file_url= ，不传时，导出到桌面
    请求体：{"data": 数据库执行接口返回的data或errors中的内容}
    """
    file_url = request.GET.get("file_url")
    file_data = json.loads(request.body).get("data")
    if not isinstance(file_data, list):
        return HttpResponse("文件数据格式为[{'data': [{}, {}]},]!")
    wb = Workbook()
    for file_data_i in file_data:
        if "data" in file_data_i and file_data_i["sql"].lower().startswith("select"):
            table_name = re.match(r'.* from (.*?)[\s|;]', file_data_i["sql"], re.I).group(1)
            ws = wb.create_sheet(table_name)
            for data_i in range(len(file_data_i["data"][0].keys())):
                ws.cell(row=1, column=data_i+1, value=list(file_data_i["data"][0].keys())[data_i])
            for data_i in range(len(file_data_i["data"])):
                for data_i_i in range(len(file_data_i["data"][data_i].values())):
                    ws.cell(row=data_i+2, column=data_i_i+1, value=list(file_data_i["data"][data_i].values())[data_i_i])
    if len(wb.sheetnames) > 1:
        del wb["Sheet"]
    if not file_url:
        # 如果没有给路径，直接导出到页面
        file_url = os.path.join(os.path.expanduser("~"), "Desktop")
    wb.save("%s/result.xlsx" % file_url)
    return HttpResponse("OK")
